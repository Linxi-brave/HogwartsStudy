#coding=utf8
import openpyxl
from openpyxl import workbook
import sys
from collections.abc import Iterable
import os
base_path = os.getcwd()
sys.path.append(base_path)

class HandExcel:
    def load_excel(self,filepath = None):
        '''
        加载excel
        '''

        if filepath == None:
            filepath = "/case/teamkitapitest.xlsx"

        else:
            filepath = filepath
        open_excel = openpyxl.load_workbook(base_path + filepath)

        return open_excel

    def get_sheet_data(self,index=None,filepath = None):
        '''
        加载所有sheet的内容
        '''
        sheet_name = self.load_excel(filepath = filepath).sheetnames
        if index == None:
            index = 0
        data = self.load_excel(filepath = filepath)[sheet_name[index]]
        return data


    def get_cell_value(self,row,cols,filepath = None):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data(filepath = filepath).cell(row=row,column=cols).value
        return data
    
    def get_rows(self,filepath = None):
        '''
        获取行数
        '''
        row = self.get_sheet_data(filepath = filepath).max_row
        return row
    
    def get_rows_value(self,row,filepath = None):
        '''
        获取某一行的内容
        '''
        row_list = []
        for i in self.get_sheet_data(filepath = filepath)[row]:
            row_list.append(i.value)
        return row_list


    def excel_write_data(self,row,cols,value,filepath = None):
        '''
        写入数据
        '''
        wb = self.load_excel(filepath)
        # wb = self.get_sheet_data(sheetindex)
        wr = wb.active
        wr.cell(row,cols,value)
        if filepath == None:
            wb.save(base_path+"/case/teamkitapitest.xlsx")
        else:
            wb.save(base_path + filepath)


    def get_columns_value(self,key=None,filepath = None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if key==None:
            key = 'A'
        columns_list_data = self.get_sheet_data(filepath = filepath)[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self,case_id,filepath = None):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value(filepath = filepath)
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num+1
        return num
    
    def get_excel_data(self,filepath = None):
        '''
        获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows(filepath)):
            data_list.append(self.get_rows_value(i+2 ,filepath))
        
        return data_list
    
excel_data = HandExcel()

if __name__ == "__main__":
    handle = HandExcel()
    filepath_jy = "/case/jyapitest.xlsx"
    data = excel_data.get_excel_data(filepath_jy)[5]
    i = 2
    if i<6:
        excel_data.excel_write_data(i, 13, "失败", filepath=filepath_jy)
        print("true")
        i = i + 1

    # for i in range(2,20):
    #     handle.get_excel_data(filepath)
    #
    #     print(handle.get_cell_value(2,i,filepath))
    #     print(i)





